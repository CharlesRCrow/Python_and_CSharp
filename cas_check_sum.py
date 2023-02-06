import openpyxl 

# Converts data from excel file to usable form

input_file=str(input("Enter File Name: "))

if '.xlsx' not in input_file:
    input_file+='.xlsx'

wb=openpyxl.load_workbook(input_file)

print('Names of Sheets in file: ',wb.sheetnames)
sheet_name=str(input('Enter name of specific sheet as seen above: '))
sheet_name=sheet_name.replace("'",'')
sh=wb[sheet_name]

sheet_cells=[]
for rows in sh.iter_rows():
    list_cas=[]
    for cell in rows:
        list_cas.append(cell.value)
    sheet_cells.append(list_cas)  
header_cells=sheet_cells[0] 
sheet_cells=sheet_cells[1:]

print(header_cells)
  
index=int(input('Column number with CAS data (start from 0): '))
   
# Functions that determines wheter CAS number is valid and reformates number into correct format

def is_cas(cas):
    cas=list(cas)
    cas_list=''.join(ch for ch in cas if ch.isdigit())
    if len(cas_list)>10 or len(cas_list)<5:
        return False
    else:
        cas_list=[int(i) for i in cas_list  ]
    return True, cas_list

def format_cas(cas_num, check_digit):
    cas_num=''.join([str(i) for i in cas_num])
    check_digit=str(check_digit)
    first_half=cas_num[:-2]
    second_half=cas_num[-2:]
    return first_half+'-'+second_half+'-'+check_digit

def check_sum(cas):
    cas=is_cas(cas)
    if cas is False:
        return 'Not a Cas Number:'
    cas=cas[1]
    cas_num=cas[:-1]
    check_digit=cas[-1]
    rev_cas_num=cas_num[::-1]
    i=1
    total=0
    clean_cas=format_cas(cas_num,check_digit)
    for digit in rev_cas_num:
        it=i*digit
        i+=1
        total+=it
    return clean_cas,(total%10==check_digit)  

cas_check=[]

for row in sheet_cells:
    value=str(row[index])
    result=check_sum(value)
    cas_check.append(result)
   
# Inserts 'Cleaned CAS' and 'Valid CAS' columns to excel file

sh.insert_cols(index+1)
cas_title=sh.cell(row=1, column=(index+1))
cas_title.value='Cleaned CAS'

sh.insert_cols(index+2)
valid_title=sh.cell(row=1, column=(index+2))
valid_title.value='Valid CAS?'

# Converts results from check_sum function into seperate lists
cas_check=list(cas_check)
cas_write_list=[]
valid_write_list=[]

for value in cas_check:
    cas_write_list.append(value[0])

for value in cas_check:
    valid_write_list.append(value[1])    

# Writes lists into the above created columns
y=2
z=0
for x in range(len(cas_check)):
    cas_write=sh.cell(row=y, column=(index+1))
    cas_write.value=cas_write_list[z]
    valid_write=sh.cell(row=y, column=(index+2))
    valid_write.value=valid_write_list[z]
    y+=1
    z+=1

wb.save(input_file)

